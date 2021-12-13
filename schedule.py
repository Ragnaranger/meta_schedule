import numpy as np
from openpyxl import Workbook
import pyswarms as ps
from pyswarms.utils.plotters import plot_cost_history
import time
import matplotlib.pyplot as plt
import openpyxl

class Teacher:
    def __init__(self, name:str, diciplines:set, array_of_restrictions:np.ndarray):
        self.list_of_constraints = []
        self.name = name

        # Disciplinas não serão implementadas agora, por enquanto, se um professor
        #  leciona mais de uma disciplina, basta adicionar ele mais de uma vez, com um nome diferente
        #  Sim, pode acontecer dele cair em horários conflitantes... (shhhhhhh...)
        self.diciplines = diciplines

        # O vetor de restrições vem no seguinte formato
        # O tamanho do vetor é o número de horários de cada dia * a quantidade de dias
        # [dia1_horario1, dia1_horario2, dia1_horario3, dia2_horario1, ...]
        # 0 é horário livre, 1 é horário proibido
        self.restrictions = array_of_restrictions

class Group:
    # O dict é no formato {disciplina:quantidade_de_aulas_necessárias}
    def __init__(self, name:str, diciplines:dict):
        self.name = name

        # Por enquanto, disciplinas tem como chave o nome do professor
        self.diciplines = diciplines


class Schedule:
    def __init__(self, n_days_in_week=5, n_classes_per_day=4):
        self.n_days_in_week = n_days_in_week
        self.n_classes_per_day = n_classes_per_day

        self.dimensions = 0

        self.list_of_teachers = []

        # Implementar turmas
        self.list_of_groups = []

    @classmethod
    def from_xlsx(cls, filename:str):
        x_file = openpyxl.load_workbook(filename)
        sheet = x_file[x_file.sheetnames[0]]
        # sheet = x_file.active
        n_dias = int(sheet['A2'].value)
        n_horas = int(sheet['B2'].value)

        s = cls(n_dias, n_horas)

        # Para cada professor na lista de professores, pegar o nome e schedule dele
        for row in sheet.iter_rows(min_row=4):
            prof_name = row[0].value
            prof_schedule = [int(cell.value) for cell in row[1:]]
            s.add_teacher(Teacher(prof_name, set(), np.array(prof_schedule)))


        # Para cada grupo na lista de grupos, pegar a quantidade de aulas que cada professor deve dar
        sheet = x_file[x_file.sheetnames[1]]
        names = sheet['A']
        for col in sheet.iter_cols(min_col=2):
            g_name = col[1].value
            disciplines = {k.value:v.value for k, v in zip(names[2:], col[2:])}
            s.add_group(Group(g_name, disciplines))

        return s


    # Compila o schedule para prepará-lo para otimização
    # Aqui, algumas variáveis são preparadas para acelerar o processo de avaliação de horários
    # Exemplo, uma matriz com a disponibilidade de professores é montada para que todos sejam comparados
    #  de uma vez
    def compile(self, n_particles, n_threads=1):
        # Quantidade de dimensões necessárias para o PSO
        self.dimensions = int(len(self.list_of_groups) * len(self.list_of_teachers) * self.n_days_in_week * self.n_classes_per_day)
        self.shape = [len(self.list_of_teachers), int(self.n_days_in_week * self.n_classes_per_day), len(self.list_of_groups)]
        self.n_particles = n_particles

        # Esse shape inclui a quantidade de partículas
        self.shape_evaluate = [n_particles//n_threads, len(self.list_of_teachers), int(self.n_days_in_week * self.n_classes_per_day), len(self.list_of_groups)]


        # Criando uma matriz com as restrições de cada professor
        all_restrictions = []
        for teacher in self.list_of_teachers:
            teacher:Teacher
            all_restrictions.append(teacher.restrictions)

        self.all_restrictions = np.array(all_restrictions)




        # Criando uma matriz com a necessidade de aulas de cada grupo
        all_group_requirements = []
        # Para cada grupo
        for group in self.list_of_groups:
            group:Group
            single_group_requirements = []

            # Verificar se existe uma necessidade com cada professor
            for teacher in self.list_of_teachers:

                # Se estiver na lista de necessidades, adicionar a quantidade especificada
                if teacher.name in group.diciplines:
                    single_group_requirements.append(group.diciplines[teacher.name])
                
                # Se não, adicionar 0
                else:
                    single_group_requirements.append(0)

                
            all_group_requirements.append(single_group_requirements)
        self.all_group_requirements = np.array(all_group_requirements).T


    def evaluate(self, x:np.ndarray):
        # Coeficiente de coeficientes que envolvem solução válida
        # Ele multiplica apenas valores que influenciam validade, não qualidade
        META_COEF_VALID_ANS = 100

        # Coeficiente de horário disponível de professor
        COEF_PROF_HR = 7 * META_COEF_VALID_ANS

        # Coeficiente de aulas que grupo deixou de receber
        COEF_GROUP_CLASS_MISSING = 5 * META_COEF_VALID_ANS

        # Coeficiente de aulas que grupo recebeu a mais
        COEF_GROUP_CLASS_EXTRA = 3 * META_COEF_VALID_ANS

        # Coeficiente professor dando mais de 1 aula por horário
        COEF_PROF_OVERLAP_TIME = 10 * META_COEF_VALID_ANS

        # Coeficiente grupo recebendo mais de 1 aula por horário
        COEF_GROUP_OVERLAP_TIME = 10 * META_COEF_VALID_ANS
        

        ##### Coeficientes de otimização de solução ####
        # Esses coeficientes não influenciam na validade da solução, apenas
        # em sua qualidade:

        # Coeficiente professor tendo dias livres
        COEF_PROF_FREE_DAY = -1

        # Coeficiente grupo tendo dias livres
        COEF_GROUP_FREE_DAY = -1

        x = x.reshape(self.shape_evaluate) # particles:professor:dia/horario:turma
        n_particles = self.shape_evaluate[0]
        points = np.zeros(self.shape_evaluate[0])

        #particle:prof:hr:turma
        # Verificar se prof estão em horários disponíveis
        prof_hora = np.sum(x, axis=3) # axis=grupo      # Atenção, não modificar prof_hora em lugar nenhum do código
        points += np.sum((prof_hora * self.all_restrictions).reshape(n_particles, -1), axis=1 ) * COEF_PROF_HR
                


        # Verificar se grupos tem as aulas que precisam
        group_aula = np.sum(x, axis=2) # axis=horario

        classes = group_aula - self.all_group_requirements
        classes = np.where(classes < 0, # Os negativos são classes faltando
                    classes * -COEF_GROUP_CLASS_MISSING, # Quando faltam aulas
                    classes * COEF_GROUP_CLASS_EXTRA) # Quando sobram aulas

        points += np.sum(classes.reshape(n_particles, -1), axis=1)




        # Verificar se professores tem horários conflitantes
        #particle:prof:hr:turma
        points += np.count_nonzero(np.sum(x, axis=(3)) > 1, axis=(1, 2)) * COEF_PROF_OVERLAP_TIME
        

        # Verificar se grupos tem horários conflitantes
        points += np.count_nonzero(np.sum(x, axis=(1)) > 1, axis=(1, 2)) * COEF_GROUP_OVERLAP_TIME



        # Verificar professores com dias livres
        # Dá 1 ponto para cada dia livre de cada professor
        # Trocar shape para particula:prof:dia:hor
        prof_dia_hora = prof_hora.reshape([n_particles, len(self.list_of_teachers), self.n_days_in_week, self.n_classes_per_day])
        points += np.count_nonzero(np.sum(prof_dia_hora, axis=3) == 0, axis=(1,2)) * COEF_PROF_FREE_DAY


        # Verificar grupos com dias livres
        # Dá 1 ponto para cada dia livre de grupo
        dia_hora_grupo = np.sum(x, axis=1).reshape(n_particles, self.n_days_in_week, self.n_classes_per_day, len(self.list_of_groups))
        points += np.count_nonzero(np.sum(dia_hora_grupo, axis=2) == 0, axis=(1,2)) * COEF_GROUP_FREE_DAY


        return points

    def sollution_to_xlsx(self, sollution:np.ndarray):

        sollution = self.decodify(sollution)

        n_profs = self.shape[0]
        n_turmas = self.shape[2]

        wb = Workbook()
        wb.remove(wb.active)

        # Para cada dia
        for day_nbr in range(self.n_days_in_week):
            sheet = wb.create_sheet('day '+ str(day_nbr))

            # Para cada turma
            for turma_nbr in range(n_turmas):
                name = self.list_of_groups[turma_nbr].name
                _ = sheet.cell(1, turma_nbr+1, name)
                # Para cada horário
                for class_nbr in range(self.n_classes_per_day):
                    index = day_nbr * self.n_classes_per_day + class_nbr

                    prof_id = np.argmax(sollution[:, index, turma_nbr])
                    
                    if sollution[prof_id, index, turma_nbr] == 1:
                        value = self.list_of_teachers[prof_id].name
                    else:
                        value = '-'

                    _ = sheet.cell(class_nbr+2, turma_nbr+1, value)
        

        wb.save('saida.xlsx')
        pass

    
    def decodify(self, x:np.ndarray):
        # professor:dia/horario:turma
        return x.reshape(self.shape)


    def add_teacher(self, teacher:Teacher):
        self.list_of_teachers.append(teacher)
    
    def add_group(self, group:Group):
        self.list_of_groups.append(group)



def example():

    s = Schedule(5, 4)
    s.add_teacher(Teacher('Daniel',   set(), np.array([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])))
    s.add_teacher(Teacher('Henrique', set(), np.array([0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1])))
    s.add_teacher(Teacher('Luiz',     set(), np.array([1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1])))
    s.add_teacher(Teacher('Gustavo',  set(), np.array([1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1])))


    s.add_teacher(Teacher('Inutil',  set(), np.array ([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])))
    s.add_teacher(Teacher('Ocupado',  set(), np.array([1, 1, 1, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1])))
    

    s.add_group(Group('Primeiro ano', {'Daniel':1, 'Henrique':1, 'Luiz':1, 'Gustavo':1}))
    s.add_group(Group('Segundo ano', {'Daniel':1, 'Henrique':1, 'Ocupado':1}))
    s.add_group(Group('Terceiro ano', {'Henrique':1, 'Luiz':1}))
    s.add_group(Group('Quarto ano', {'Inutil':1, 'Ocupado':1, 'Daniel':1}))

    # n_particles precisa ser múltiplo de n_threads
    n_particles = 40
    n_threads = 1
    s.compile(n_particles, n_threads)


    options = {'c1': 100000000.0, 'c2': 0.6, 'w':0.9, 'k':4, 'p':1}

    start = time.time()
    optmizer = ps.discrete.BinaryPSO(n_particles, s.dimensions, options=options)
    cost, pos = optmizer.optimize(s.evaluate, iters=10000, n_processes=n_threads)
    # print('Cost: ', cost,'Pos: ', pos)
    # print(pos.reshape(s.shape))
    stop = time.time()
    print('Duration: ', stop-start)

    plot_cost_history(optmizer.cost_history)
    plt.show()



if __name__ == '__main__':
    # test()
    example()